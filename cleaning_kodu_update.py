# Temizlik kodu
# Bu kod wpdeki excel alıp temizler, sütunları düzeltir küfürleri ve ingilizce ve saçma kelimeler içeren satırları siler.
# Genel Duygu Analizi katmanı eklendi.

import pandas as pd
import re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
#from langdetect import detect # Bu satır kullanılmıyorsa kaldırılabilir
import json
from zemberek import (
    TurkishMorphology,
)
from openpyxl.styles import PatternFill
import time # Zamanlama için eklendi (opsiyonel)

# --- Dosya Yolları ve Sabitler ---
INPUT_PATH = r"C:\Users\alibaki.turkoz\Desktop\ÖZÜ_CS_MASTER\Semester_2\CS549_Introduction_to_Natural_Language_Processing\Group_Project\Codes\dataset\reviews.csv"
OUTPUT_PATH = r"C:\Users\alibaki.turkoz\Desktop\ÖZÜ_CS_MASTER\Semester_2\CS549_Introduction_to_Natural_Language_Processing\Group_Project\Codes\dataset\cleaned_reviews_zemberek_GUNCEL.xlsx"

# KEYWORDS
ASPECTS_KEYWORDS_PATH = r"C:\Users\alibaki.turkoz\Desktop\ÖZÜ_CS_MASTER\Semester_2\CS549_Introduction_to_Natural_Language_Processing\Group_Project\Codes\data_preparation\constants\aspects_keywords_guncel.json"
SENTIMENT_KEYWORDS_PATH = r"C:\Users\alibaki.turkoz\Desktop\ÖZÜ_CS_MASTER\Semester_2\CS549_Introduction_to_Natural_Language_Processing\Group_Project\Codes\data_preparation\constants\sentiment_keywords_guncel.json"
# COLORS_MAP_PATH = "./constants/colors_map.py" # Doğrudan tanımlandığı için kaldırıldı

# Initialize Zemberek Morphology
print("Initializing Zemberek Morphology...")
start_time = time.time()
morphology = TurkishMorphology.create_with_defaults()
end_time = time.time()
print(f"Zemberek Morphology initialized in {end_time - start_time:.4f} seconds.")

# Keyword lists will be loaded from JSON
ASPECTS_KEYWORDS = {}
SENTIMENT_KEYWORDS = {}

# Renk Haritası
COLOR_MAP = {
    "Olumlu": PatternFill(start_color="90EE90", fill_type="solid"), # Açık Yeşil
    "Olumsuz": PatternFill(start_color="FFA07A", fill_type="solid"), # Açık Kırmızı/Somon
    "Nötr": PatternFill(start_color="D3D3D3", fill_type="solid"),   # Açık Gri
}

# --- Yardımcı Fonksiyonlar ---

# read json file
def read_json(file_path):
    try:
        with open(file_path, "r", encoding='utf-8') as f:
            data = json.load(f)
        return data
    except FileNotFoundError:
        print(f"Error: JSON file not found at {file_path}")
        return None
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from {file_path}")
        return None

# Convert any given text to string and lowercase
def to_str_lowercase(text):
    return str(text).lower()

# Replace Turkish characters with English equivalents for normalization
def normalize_text(text):
    replacements = str.maketrans("ÇĞİÖŞÜçğıöşü", "CGIOSUcgiosu")
    return text.translate(replacements)

# Normalize a single keyword (used for keyword lists before lemmatization)
def normalize_keyword_simple(keyword):
    keyword = to_str_lowercase(keyword)
    keyword = normalize_text(keyword)
    return keyword

# Remove extra spaces from the text
def remove_extra_spaces(text):
    return re.sub(r'\s+', ' ', text).strip()

# --- ZEMBEREK BASED LEMMATIZATION AND MATCHING ---
# Function to get lemma of a word using Zemberek
def get_lemma(word):
    try:
        # Kelimeyi küçük harfe çevirip analiz et
        analysis_results = morphology.analyze(word.lower())
        # Zemberek bazen boş liste veya None dönebilir
        if analysis_results and isinstance(analysis_results, list) and len(analysis_results) > 0:
            # En olası analizi al (genellikle ilk sonuç)
            best_analysis = analysis_results[0]
            # Analiz sonucundan lemmayı al
            lemma = best_analysis.get_lemma()
            # Eğer lemma boş değilse döndür
            if lemma:
                return lemma
    except Exception as e:
        # Hata durumunda loglama yapılabilir (opsiyonel)
        # print(f"Error analyzing word '{word}': {e}")
        pass # Hata durumunda orijinal kelimeyi döndürmeye devam et
    # Analiz başarısız olursa veya lemma bulunamazsa orijinal kelimeyi (küçük harfle) döndür
    return word.lower()


# Pre-lemmatize and normalize a text, returning a set of unique lemmas
def lemmatize_normalize_text(text):
    text = to_str_lowercase(text) # Lowercase the input text
    # Tokenize text into words (Türkçe karakterleri de içeren kelimeleri bul)
    words = re.findall(r'\b[a-zA-ZÇĞİÖŞÜçğıöşü]+\b', text)
    # Lemmatize each word, normalize (English chars), and store unique lemmas in a set
    lemmatized_normalized_words = {normalize_text(get_lemma(word)) for word in words if len(word) > 1} # Tek harfli kelimeleri atla
    return lemmatized_normalized_words

# Check if a normalized keyword's lemma exists in the lemmatized text set
def has_lemmatized_keyword(text_lemmatized_normalized_set, normalized_keyword_lemma):
    return normalized_keyword_lemma in text_lemmatized_normalized_set

# --- PROCESS CONSTANTS (with Zemberek Lemmatization) ---

# Lemmatize and normalize all keywords in the dictionaries
def lemmatize_normalize_keywords(keywords_dict):
    processed_dict = {}
    if keywords_dict is None:
        print("Warning: Keyword dictionary is None. Returning empty dictionary.")
        return processed_dict
    for category, keywords in keywords_dict.items():
        if not isinstance(keywords, list):
            print(f"Warning: Keywords for category '{category}' is not a list. Skipping.")
            continue
        # Lemmatize and normalize keywords to lowercase and English chars
        # Ensure keywords are strings before processing
        lemmatized_normalized_keywords = {normalize_text(get_lemma(str(k))) for k in keywords if k} # Boş keywordleri atla
        processed_dict[category] = lemmatized_normalized_keywords # Use set for faster lookups
    return processed_dict

# Load keywords from JSON
aspects_data = read_json(ASPECTS_KEYWORDS_PATH)
sentiments_data = read_json(SENTIMENT_KEYWORDS_PATH)

# Check if data loaded successfully before accessing keys
if aspects_data and "ASPECTS_KEYWORDS" in aspects_data:
    ASPECTS_KEYWORDS = aspects_data["ASPECTS_KEYWORDS"]
else:
    print("Error: Could not load or find ASPECTS_KEYWORDS in the JSON file.")
    ASPECTS_KEYWORDS = {} # Boş bırak veya hata ver

if sentiments_data and "SENTIMENT_KEYWORDS" in sentiments_data:
    SENTIMENT_KEYWORDS = sentiments_data["SENTIMENT_KEYWORDS"]
else:
    print("Error: Could not load or find SENTIMENT_KEYWORDS in the JSON file.")
    SENTIMENT_KEYWORDS = {} # Boş bırak veya hata ver


# Lemmatize and normalize keywords once at the start
print("Lemmatizing and normalizing keywords...")
start_time = time.time()
LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS = lemmatize_normalize_keywords(ASPECTS_KEYWORDS)
LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS = lemmatize_normalize_keywords(SENTIMENT_KEYWORDS)
end_time = time.time()
print(f"Keywords processed in {end_time - start_time:.4f} seconds.")

# Calculate and print keyword counts
total_aspect_keywords = sum(len(keywords) for keywords in LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS.values())
total_pos_sentiment_keywords = len(LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS.get("Olumlu", set()))
total_neg_sentiment_keywords = len(LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS.get("Olumsuz", set()))
print(f"Lemmatized Aspects keywords count: {total_aspect_keywords}")
print(f"Lemmatized Sentiment keywords count: {total_pos_sentiment_keywords + total_neg_sentiment_keywords} (Olumlu: {total_pos_sentiment_keywords}, Olumsuz: {total_neg_sentiment_keywords})")

# Example: Print some lemmatized keywords to verify
if 'Grafik' in LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS:
    print("Sample lemmatized 'Grafik' keywords: ", list(LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS['Grafik'])[:5])
if 'Olumlu' in LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS:
    print("Sample lemmatized 'Olumlu' keywords: ", list(LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS['Olumlu'])[:5])


# ---- MAIN ANALYSIS FUNCTION (Updated with Zemberek & Sentence-Level Sentiment & Genel Duygu) ----
# Define common Turkish negation word lemmas (adjust if Zemberek produces different lemmas)
# Lemmatize negation words as well
NEGATION_LEMMAS = {normalize_text(get_lemma(word)) for word in ["değil", "yok", "hiç", "asla", "olumsuz", "kötü"]} # Normalize edilmiş lemmatize halleri

def analyze_review(text):
    # Initialize result with Nötr for all aspects AND Genel Duygu
    result = {aspect: "Nötr" for aspect in LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS}
    result["Genel Duygu"] = "Nötr" # YENİ EKLENDİ

    # Ensure text is a string before processing
    if pd.isna(text):
        return result # Return default neutral if input is NaN
    original_text_lower = str(text).lower() # Keep original lowercased text for sentence splitting

    # Lemmatize and normalize the entire review text once for efficiency
    review_lemmatized_normalized_set = lemmatize_normalize_text(original_text_lower)
    if not review_lemmatized_normalized_set: # Eğer lemmatize edilecek kelime yoksa nötr dön
        return result

    found_aspects_with_keywords = {} # Store aspect and the specific keyword lemma found

    # First pass: Identify all aspects mentioned
    for aspect, lemmatized_keywords_set in LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS.items():
        # Use intersection for efficiency
        matching_keyword_lemmas = review_lemmatized_normalized_set.intersection(lemmatized_keywords_set)
        if matching_keyword_lemmas:
            # Store the aspect name if any keyword lemma matches
            found_aspects_with_keywords[aspect] = True # Sadece bulunduğunu işaretlemek yeterli

    # Second pass: Determine sentiment for found aspects based on relevant sentences
    aspect_sentiment_assigned = False # Flag to check if any aspect got a non-neutral sentiment
    if found_aspects_with_keywords:
        # Split original lowercased text into sentences
        sentences = re.split(r'[.!?]+', original_text_lower)
        pos_keywords_lemmas = LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS.get("Olumlu", set())
        neg_keywords_lemmas = LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS.get("Olumsuz", set())

        for aspect in found_aspects_with_keywords: # Iterate through aspects found in the review
            aspect_sentiment_score = 0
            # Get all lemmatized & normalized keyword lemmas for the current aspect
            aspect_keywords_set = LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS[aspect]

            for sentence in sentences:
                sentence = sentence.strip()
                if not sentence: continue # Skip empty sentences

                # Lemmatize and normalize the current sentence
                sentence_lemmatized_normalized_set = lemmatize_normalize_text(sentence)
                if not sentence_lemmatized_normalized_set: continue # Skip if sentence has no processable words

                # Check if the sentence contains any keyword lemma for the current aspect
                # Use isdisjoint for faster check (True if no common elements)
                if not aspect_keywords_set.isdisjoint(sentence_lemmatized_normalized_set):
                    # Sentence is relevant to the aspect, now analyze its sentiment
                    positive_matches = sentence_lemmatized_normalized_set.intersection(pos_keywords_lemmas)
                    negative_matches = sentence_lemmatized_normalized_set.intersection(neg_keywords_lemmas)

                    sentence_score = len(positive_matches) - len(negative_matches)

                    # Basic Negation Check for the sentence: If a negation lemma exists, neutralize the score
                    if not NEGATION_LEMMAS.isdisjoint(sentence_lemmatized_normalized_set):
                        sentence_score = 0 # Neutralize sentence score if negation found

                    aspect_sentiment_score += sentence_score

            # Determine final sentiment for the aspect based on aggregated score
            if aspect_sentiment_score > 0:
                result[aspect] = "Olumlu"
                aspect_sentiment_assigned = True # Mark that sentiment was assigned
            elif aspect_sentiment_score < 0:
                result[aspect] = "Olumsuz"
                aspect_sentiment_assigned = True # Mark that sentiment was assigned
            # else: remains "Nötr" (if score is 0)

    # --- YENİ BÖLÜM: Genel Duygu Analizi ---
    # Eğer hiçbir aspect'e özel bir duygu atanmadıysa VE yorumda genel duygu kelimeleri varsa
    if not aspect_sentiment_assigned:
        pos_keywords_lemmas = LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS.get("Olumlu", set())
        neg_keywords_lemmas = LEMMATIZED_NORMALIZED_SENTIMENT_KEYWORDS.get("Olumsuz", set())

        # Tüm review'daki sentiment kelimelerini bul (zaten hesaplanmıştı)
        positive_matches_overall = review_lemmatized_normalized_set.intersection(pos_keywords_lemmas)
        negative_matches_overall = review_lemmatized_normalized_set.intersection(neg_keywords_lemmas)

        # Sadece duygu kelimesi varsa genel duygu ata
        if positive_matches_overall or negative_matches_overall:
            overall_sentiment_score = len(positive_matches_overall) - len(negative_matches_overall)

            # Tüm review için genel negation kontrolü (basit yaklaşım)
            # Eğer yorumun tamamında olumsuzlama varsa, skoru nötrle
            if not NEGATION_LEMMAS.isdisjoint(review_lemmatized_normalized_set):
                 overall_sentiment_score = 0 # Genel skoru nötrle

            # Genel duygu sonucunu ata
            if overall_sentiment_score > 0:
                result["Genel Duygu"] = "Olumlu"
            elif overall_sentiment_score < 0:
                result["Genel Duygu"] = "Olumsuz"
            # else: remains "Nötr" (if score is 0 or neutralized)

    return result

# --- VERİ YÜKLE ---
print(f"Loading reviews from {INPUT_PATH}...")
try:
    # Sadece gerekli sütunları oku ve metin olmayanları atla
    df = pd.read_csv(INPUT_PATH, usecols=["app_id", "review_text"], dtype={"app_id": str, "review_text": str})
    df = df.dropna(subset=['review_text'])
    df = df[df['review_text'].str.strip() != ''] # Boş yorumları da atla
    print(f"Successfully loaded {len(df)} non-empty reviews.")
    print("Original DataFrame head:")
    print(df.head())
except FileNotFoundError:
    print(f"Error: Input file not found at {INPUT_PATH}. Please check the path.")
    exit() # Dosya yoksa çık
except Exception as e:
    print(f"An error occurred while loading the CSV: {e}")
    exit() # Başka bir hata olursa çık


# --- ANALİZ (Using Zemberek & Genel Duygu) ---
print("\nStarting analysis...")
results = []
total_rows = len(df)
analysis_start_time = time.time()

print(f"Processing {total_rows} reviews...")
for index, row in df.iterrows():
    # review_text'in string olduğundan emin ol (nadiren de olsa farklı tip gelebilir)
    review_text = str(row["review_text"]) if pd.notna(row["review_text"]) else ""

    if not review_text: # Eğer yorum boşsa atla (dropna sonrası tekrar kontrol)
        analysis = {aspect: "Nötr" for aspect in LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS}
        analysis["Genel Duygu"] = "Nötr"
    else:
        analysis = analyze_review(review_text)

    results.append({
        "app_id": row["app_id"],
        "review_text": review_text, # Orijinal metni sakla
        **analysis # Analiz sonuçlarını ekle (aspectler + Genel Duygu)
    })

    # Print progress
    if (index + 1) % 500 == 0 or (index + 1) == total_rows:
        elapsed_time = time.time() - analysis_start_time
        print(f"  Processed {index + 1}/{total_rows} reviews... (Elapsed: {elapsed_time:.2f}s)")

df_results = pd.DataFrame(results)
analysis_end_time = time.time()
print(f"\nAnalysis complete. Processed {len(df_results)} reviews in {analysis_end_time - analysis_start_time:.2f} seconds.")
print("Results DataFrame head:")
print(df_results.head())


# --- RENKLİ KAYDET (Updated for Zemberek & Genel Duygu) ---
print(f"\nSaving results to {OUTPUT_PATH}...")
save_start_time = time.time()
try:
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df_results.to_excel(writer, index=False, sheet_name="Analizler")
        worksheet = writer.sheets["Analizler"]

        # Renklendirilecek sütunları belirle (Aspectler + Genel Duygu)
        # Sadece df_results'da gerçekten var olan sütunları al
        columns_to_color = [col for col in list(LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS.keys()) + ["Genel Duygu"] if col in df_results.columns]

        color_col_indices = {}
        for col_name in columns_to_color:
             # +1 for 1-based index in openpyxl
            color_col_indices[col_name] = df_results.columns.get_loc(col_name) + 1

        # Apply coloring to specified columns
        print("Applying colors to Excel sheet...")
        for col_name, col_idx in color_col_indices.items():
            col_letter = get_column_letter(col_idx)
            # print(f"  Coloring column: {col_name} ({col_letter})") # Detaylı loglama için açılabilir
            for row in range(2, len(df_results) + 2): # +2 because Excel is 1-based and header is row 1
                cell = worksheet[f"{col_letter}{row}"]
                sentiment_value = cell.value
                # Use get() for safer access to COLOR_MAP, default to Nötr color
                cell.fill = COLOR_MAP.get(sentiment_value, COLOR_MAP["Nötr"])

        # Adjust column widths (optional)
        print("Adjusting column widths...")
        worksheet.column_dimensions['A'].width = 15 # app_id
        worksheet.column_dimensions['B'].width = 80 # review_text
        for col_name, col_idx in color_col_indices.items():
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 20 # Aspect ve Genel Duygu sütunları

    save_end_time = time.time()
    print(f"Main analysis sheet saved and formatted in {save_end_time - save_start_time:.2f} seconds.")

    # --- ÖZET SHEET (Updated for Zemberek & Genel Duygu) ---
    print("Creating summary sheet...")
    summary_start_time = time.time()
    summary_data = []
    # Özetlenecek kategoriler (Aspectler + Genel Duygu) - Sadece df_results'da olanlar
    categories_for_summary = [col for col in list(LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS.keys()) + ["Genel Duygu"] if col in df_results.columns]

    for category in categories_for_summary:
        counts = df_results[category].value_counts().to_dict()
        summary_data.append({
            "Kategori": category,
            "Olumlu": counts.get("Olumlu", 0),
            "Olumsuz": counts.get("Olumsuz", 0),
            "Nötr": counts.get("Nötr", 0)
        })
    df_summary = pd.DataFrame(summary_data)

    # Append summary sheet to the existing Excel file
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl", mode="a") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Özet")
    summary_end_time = time.time()
    print(f"Summary sheet created in {summary_end_time - summary_start_time:.2f} seconds.")


    # --- NÖTR KELİMELER (Updated for Zemberek & Genel Duygu) ---
    print("Finding most common words in fully neutral reviews...")
    neutral_words_start_time = time.time()
    # Dynamically create the filter condition for all aspects AND Genel Duygu being Nötr
    neutral_condition = None
    # Kontrol edilecek sütunlar (Aspectler + Genel Duygu) - Sadece df_results'da olanlar
    columns_for_neutral_check = [col for col in list(LEMMATIZED_NORMALIZED_ASPECTS_KEYWORDS.keys()) + ["Genel Duygu"] if col in df_results.columns]

    # Build the combined condition using boolean indexing
    if columns_for_neutral_check: # Eğer kontrol edilecek sütun varsa
        neutral_conditions = [(df_results[col] == "Nötr") for col in columns_for_neutral_check]
        # Combine all conditions with logical AND
        combined_neutral_condition = pd.concat(neutral_conditions, axis=1).all(axis=1)
        neutral_df = df_results[combined_neutral_condition]
    else:
        neutral_df = pd.DataFrame() # Boş DataFrame

    if not neutral_df.empty:
        neutral_reviews = neutral_df["review_text"].tolist()
        print(f"Found {len(neutral_reviews)} fully neutral reviews.")

        # Join all neutral reviews into a single corpus
        neutral_corpus = " ".join(map(str, neutral_reviews)).lower()

        # Re-tokenize and lemmatize the corpus for accurate counting
        words_for_counting = re.findall(r'\b[a-zA-ZÇĞİÖŞÜçğıöşü]+\b', neutral_corpus)
        # Lemmatize, normalize, and filter short words
        lemmatized_words_for_counting = [
            normalize_text(get_lemma(word))
            for word in words_for_counting
            if len(word) >= 3 # En az 3 harfli kelimeleri say
        ]

        # Count word frequencies
        if lemmatized_words_for_counting:
            most_common_words = Counter(lemmatized_words_for_counting).most_common(100)
            df_top_words = pd.DataFrame(most_common_words, columns=["Kelime (Lemma)", "Frekans"])

            # Append the top words sheet
            with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl", mode="a") as writer:
                df_top_words.to_excel(writer, index=False, sheet_name="Nötr En Çok Kelimeler (Lemma)")
            print("Neutral words sheet created.")
        else:
             print("⚠️ No words found in neutral reviews after processing.")

    else:
        print("⚠️ No fully neutral reviews found or relevant columns missing for neutral word analysis.")

    neutral_words_end_time = time.time()
    print(f"Neutral words analysis completed in {neutral_words_end_time - neutral_words_start_time:.2f} seconds.")

    print(f"\nProcess Done✅: Output saved to {OUTPUT_PATH}")

except Exception as e:
    print(f"\nAn error occurred during saving or post-processing: {e}")
    print("Please check file permissions and data integrity.")

