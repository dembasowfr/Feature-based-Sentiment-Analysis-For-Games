import pandas as pd
import re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# --- YAPILANDIRMA ---
INPUT_PATH = r"C:\Users\alibaki.turkoz\Desktop\CS549_Project\Games_Rewievs_Dataset_Excel.xlsx"
OUTPUT_PATH = r"C:\Users\alibaki.turkoz\Desktop\CS549_Project\Games_Sentiment_Analysis_Results.xlsx"

ASPECT_KEYWORDS = {
    "Grafik": ["grafik", "görsel", "görüntü", "tasarım", "renk", "art", "doku", "estetik",
               "görsellik", "atmosfer", "pixel", "animasyon", "render", "vfx", "efekt",
               "gölgelendirme", "detaylı", "parlak", "canlı", "ışıklandırma", "derinlik",
               "kompozisyon", "minimal", "soyut", "modern", "klasik", "retro"],
    "AI (yapay zeka)": ["ai", "yapay zeka", "düşman", "boss", "algoritma", "öğrenme",
                        "adaptasyon", "strateji", "npc", "davranış", "yapay", "zeka",
                        "patika", "karar", "mantık", "otonom", "reaksiyon", "sistem"],
    "Oyunun Oynanışı": ["oynanış", "kontrol", "mekanik", "sistem", "combo", "hikaye", "zor", "oyun",
                        "zorluk", "akış", "kontroller", "oyunu", "oyunun", "oyunlar", "mükemmel",
                        "harika", "gerçekçi", "akıcı", "kaliteli", "tavsiye", "güzel", "iyi",
                        "büyüleyici", "en iyi", "öneririm", "strateji", "refleks", "etkileşim",
                        "deneyim", "gerçek zamanlı", "çoklu görev", "işbirliği", "rekabet", "dinamik",
                        "oyunculuk", "akıcılık", "karmaşık", "esnek", "özelleştirilebilir", "hızlı tempolu"]
}

SENTIMENT_WORDS = {
    "Olumlu": [
        "mükemmel", "harika", "gerçekçi", "akıcı", "kaliteli", "tavsiye", "güzel", "iyi", "büyüleyici", "en iyi",
        "beğeni", "üstün", "fevkalade", "olağanüstü", "göz alıcı", "muhteşem", "etkileyici", "cazip", "özenli",
        "profesyonel", "başarılı", "takdir", "memnun", "harikulade", "nefes kesici", "şaheser", "akıl almaz",
        "sürükleyici", "bağımlılık yapan", "unutulmaz", "yenilikçi", "övgü", "taktir", "alkış", "müthiş",
        "kusursuz", "ideal", "parlak", "dahice", "keyifli", "sarıyo", "sarıyor", "oynayın", "zevkli", "süper",
        "başyapıt", "aşık", "aşk", "baş yapıt", "efsane", "mük", "kalite", "taş", "inanılmaz", "hoş", "fantastik",
        "eğlenceli", "tatmin edici", "şahane", "yeterli", "yetenekli", "hayranlık verici", "doyurucu", "ustaca",
        "tatlı", "nefis", "görkemli", "zarif", "dinamik", "etkin", "rahatlatıcı", "seçkin", "aydınlık", "umut verici",
        "hoşnut edici", "kararlı", "güçlü", "becerikli", "sevindirici", "nitelikli", "canlı", "çekici", "ferah",
        "açık", "dengeli", "ilginç", "başarımı yüksek", "güvenilir", "eğitimli", "kusursuzca", "uyumlu", "verimli",
        "yetkin", "sadık", "motive edici", "pozitif", "iyi niyetli", "şık", "temiz", "basit", "anlaşılır", "özgün",
        "heyecan verici", "dikkat çekici", "baş döndürücü", "iç açıcı", "parlak fikirli", "vizyoner", "çığır açan",
        "enfes", "tadında", "yüce", "yüksek kaliteli", "iyi düşünülmüş", "akıllıca", "mükemmel tasarım",
        "tatmin edici performans", "canlandırıcı", "rahat", "modern", "enerjik", "sadelik", "doygun", "akıl dolu",
        "heyecanlı", "esprili", "olumlu hava", "minnettar", "etkileyicilik", "en üst seviye", "kolay", "pratik",
        "verimli çözüm", "en iyisi", "üst düzey", "parlak zeka", "çok iyi", "olumlu izlenim", "gelişmiş",
        "tatminkâr", "başarı", "yetkinlik", "ilerleme", "seviye atlamış", "ustalaşmış", "akıcı oynanış",
        "takdire şayan", "zirve", "öncü", "rehber", "akılcı", "hızlı", "zekice", "mutlu", "eğlencelik",
        "oynanabilir", "şanslı", "enfes görsellik", "şahane tasarım", "çok güzel", "fevkaladelik", "kapsayıcı",
        "etkinleştirici", "etkili", "çözümleyici"
    ],
    "Olumsuz": [
        "kötü", "oynamayın", "kanser", "berbat", "hatalı", "yavaş", "sıkıcı", "donuk", "zayıf", "kötücül",
        "işe yaramaz", "eksik", "yetersiz", "kaba", "acımasız", "düzensiz", "hayal kırıklığı", "verimsiz",
        "mutsuz", "çirkin", "fiyasko", "tatsız", "tekdüze", "vasat", "sinir bozucu", "tutarsız", "başarısız",
        "kırık", "çöküyor", "dengesiz", "kafa karıştırıcı", "pişman", "iğrenç", "rezil", "kalitesiz",
        "dayanılmaz", "bozuk", "çöp", "almayın", "karmaşık", "anlamsız", "kırılmış", "gereksiz", "hoş olmayan",
        "beceriksiz", "sorunlu", "sıkıntılı", "şaşırtıcı derecede kötü", "yoğun", "ağır", "yavaşlatıcı",
        "yanlış", "absürt", "çürük", "akılsızca", "rahatsız edici", "nefret edilesi", "bozucu", "negatif", "eksi",
        "kötüleşmiş", "zaman kaybı", "felaket", "değersiz", "fazla", "kararsız", "dikkatsiz", "bulanık",
        "gereksizce", "sahte", "korkunç", "ürkütücü", "rezalet", "rahatsız", "saldırgan", "uygunsuz", "sefil",
        "düşük kaliteli", "çökük", "eski", "bozulmuş", "geçersiz", "karışık", "kararsızlık", "sorun çıkartan",
        "yönsüz", "eksik içerik", "berbat hissiyat", "eksik fonksiyon", "gereksiz uzatılmış", "karmakarışık",
        "net olmayan", "boş", "uzun", "gıcık", "mantıksız", "karışıklık", "çözüm yok", "uyumsuz", "kaza", "riskli",
        "aşırı zor", "bıktırıcı", "kusurlu", "ezbere", "sinir bozucu deneyim", "gereksiz diyalog", "çok yavaş",
        "tekrarlayan", "eski teknoloji", "sistemsiz", "yenilikten uzak", "bağlantısız", "eksik veri",
        "hatalı tasarım", "başarısızlık", "sinir edici", "dengesizlik", "devamsız", "hatalı yön", "eksik bilgi",
        "yanlış yönlendirme", "eksik oyun", "donanım hatası", "yanıltıcı", "fazlalık", "zayıflık", "bozuk görsel",
        "görsel karmaşa", "anlamsız görevler", "hatalı grafik", "ses hatası", "anlamsız oyun", "gereksiz detay",
        "eksik içerik"
    ]
}

COLOR_MAP = {
    "Olumlu": PatternFill(start_color="90EE90", fill_type="solid"),
    "Olumsuz": PatternFill(start_color="FFA07A", fill_type="solid"),
    "Nötr": PatternFill(start_color="D3D3D3", fill_type="solid")
}


# --- FONKSIYONLAR ---
def normalize_text(text):
    replacements = str.maketrans("çğıöşü", "cgiosu")
    return text.translate(replacements)


def has_keyword_with_suffix(text, keyword):
    suffixes = [
        "", "i", "ı", "u", "ü", "e", "a", "de", "da", "den", "dan", "ye", "ya", "nin", "nın", "nun", "nün",
        "ne", "na", "ler", "lar", "si", "sı", "su", "sü", "inden", "undan", "lerinin", "larının", "deki",
        "daki", "yle", "yla", "li", "lı", "lu", "lü", "ik", "ık", "ük", "ek", "ak"
    ]
    keyword = normalize_text(keyword.lower())
    text = normalize_text(text.lower())
    pattern = r'\b' + f"{keyword}({'|'.join(suffixes)})?" + r'\b'
    return re.search(pattern, text) is not None


def has_extended_match(text, keyword):
    text = normalize_text(text.lower())
    keyword = normalize_text(keyword.lower())
    pattern = r'\b\w*' + re.escape(keyword) + r'\w*\b'
    matches = re.findall(pattern, text)
    for match in matches:
        if len(match) >= len(keyword) + 2:
            return True
    return keyword in text


def keyword_match(text, keyword):
    return has_keyword_with_suffix(text, keyword) or has_extended_match(text, keyword)


def detect_sentiment(text, aspect_keyword):
    try:
        text = str(text).lower()
        aspect_keyword = aspect_keyword.lower()
        sentences = re.split(r'[.!?]', text)
        target_sentences = [sentence.strip() for sentence in sentences if keyword_match(sentence, aspect_keyword)]
        if not target_sentences:
            return "Nötr"

        context_text = " ".join(target_sentences)
        words = re.findall(r'\b\w+\b', normalize_text(context_text.lower()))
        pos = [normalize_text(w) for w in SENTIMENT_WORDS["Olumlu"]]
        neg = [normalize_text(w) for w in SENTIMENT_WORDS["Olumsuz"]]
        positive = sum(1 for word in words if word in pos)
        negative = sum(1 for word in words if word in neg)

        if positive > negative:
            return "Olumlu"
        elif negative > positive:
            return "Olumsuz"
        else:
            return "Nötr"
    except Exception as e:
        print(f"⚠️ Sentiment analiz hatası: {str(e)}")
        return "Nötr"


def analyze_review(text):
    result = {aspect: "Nötr" for aspect in ASPECT_KEYWORDS}
    text = str(text).lower()
    for aspect, keywords in ASPECT_KEYWORDS.items():
        for keyword in keywords:
            if keyword_match(text, keyword):
                sentiment = detect_sentiment(text, keyword)
                if sentiment != "Nötr":
                    result[aspect] = sentiment
                    break
    return result


# --- VERİ YÜKLE ---
df = pd.read_excel(INPUT_PATH, usecols=["app_id", "review_text"], dtype={"app_id": str, "review_text": str}).dropna()

# --- ANALİZ ---
results = []
for _, row in df.iterrows():
    analysis = analyze_review(row["review_text"])
    results.append({
        "app_id": row["app_id"],
        "review_text": row["review_text"],
        **analysis
    })
df_results = pd.DataFrame(results)

# --- RENKLİ KAYDET ---
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    df_results.to_excel(writer, index=False, sheet_name="Analizler")
    worksheet = writer.sheets["Analizler"]
    for col in range(3, 6):
        col_letter = get_column_letter(col)
        for row in range(2, len(df_results) + 2):
            cell = worksheet[f"{col_letter}{row}"]
            cell.fill = COLOR_MAP.get(cell.value, COLOR_MAP["Nötr"])
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 80

# --- ÖZET SHEET ---
summary_data = []
for aspect in ASPECT_KEYWORDS:
    counts = df_results[aspect].value_counts().to_dict()
    summary_data.append({
        "Kategori": aspect,
        "Olumlu": counts.get("Olumlu", 0),
        "Olumsuz": counts.get("Olumsuz", 0),
        "Nötr": counts.get("Nötr", 0)
    })
df_summary = pd.DataFrame(summary_data)
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl", mode="a") as writer:
    df_summary.to_excel(writer, index=False, sheet_name="Özet")

# --- NÖTR KELİMELER ---
neutral_reviews = df_results[
    (df_results["Grafik"] == "Nötr") &
    (df_results["AI (yapay zeka)"] == "Nötr") &
    (df_results["Oyunun Oynanışı"] == "Nötr")
]["review_text"].tolist()
neutral_corpus = " ".join(map(str, neutral_reviews)).lower()
neutral_corpus = normalize_text(neutral_corpus)
neutral_words = re.findall(r'\b\w{3,}\b', neutral_corpus)
most_common_words = Counter(neutral_words).most_common(100)
df_top_words = pd.DataFrame(most_common_words, columns=["Kelime", "Frekans"])
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl", mode="a") as writer:
    df_top_words.to_excel(writer, index=False, sheet_name="Nötr En Çok Kelimeler")

print(f"✅ İşlem tamamlandı! Excel çıktısı: {OUTPUT_PATH}")
